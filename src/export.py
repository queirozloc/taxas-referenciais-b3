import io
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_DATE_FILL = PatternFill("solid", fgColor="D6E4F0")
_DATE_FONT = Font(bold=True, size=10)
_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_RATE_FORMAT = "0.0000"


def export_to_excel(
    di_df: pd.DataFrame,
    cupom_df: pd.DataFrame,
    output_path: str | io.BytesIO,
) -> None:
    """Write DI and Cupom Cambial Limpo curves to a formatted Excel workbook."""
    with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="DD/MM/YYYY") as writer:
        _write_sheet(writer, di_df, "DI Curve")
        _write_sheet(writer, cupom_df, "Cupom Cambial Limpo")
    if isinstance(output_path, str):
        print(f"Saved: {output_path}")


def _write_sheet(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    pivot = (
        df.pivot(index="date", columns="tenor", values="rate")
        .sort_index()
    )
    pivot.index = pd.to_datetime(pivot.index)

    pivot.to_excel(writer, sheet_name=sheet_name)

    ws = writer.sheets[sheet_name]
    _format_sheet(ws, n_cols=pivot.shape[1])


def _format_sheet(ws, n_cols: int) -> None:
    # Header row (row 1): tenor labels
    for col_idx in range(1, n_cols + 2):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    # Data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for i, cell in enumerate(row):
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
            if i == 0:
                # Date column
                cell.fill = _DATE_FILL
                cell.font = _DATE_FONT
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.number_format = _RATE_FORMAT

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 14  # date column
    for col_idx in range(2, n_cols + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11

    ws.freeze_panes = "B2"
